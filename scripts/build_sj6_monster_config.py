#!/usr/bin/env python3
"""Build a CoursewareMaker 贪吃小怪兽 config from an Excel sheet.

The script is intentionally deterministic:
- input paths are explicit CLI parameters;
- resource names must be unique and category-correct;
- every level must have exactly 3 options and exactly 1 correct option;
- prompt image/text fields must be written to a matching template component.
"""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from copy import deepcopy
from pathlib import Path
from typing import Any
from uuid import uuid4

import openpyxl


DEFAULT_RESOURCE_JSON = Path("resources/latest_resources.json")
DEFAULT_QUESTION_XLSX = Path("zhiyinlou_monster_test.xlsx")
DEFAULT_TEMPLATE_JSON = Path("reference_configs/monster/贪吃_reference_clean.json")
DEFAULT_OUTPUT_JSON = Path("Sj6贪吃小怪兽.config.json")
DEFAULT_OUTPUT_META = Path("Sj6贪吃小怪兽.build-meta.json")
DEFAULT_SHEET_NAME = "Sj6贪吃小怪兽"

NODE_NAME_BY_OPTION = {1: "节点", 2: "节点_104", 3: "节点_103"}
RIGHT_ANIMATION_BY_OPTION = {1: "right_1_2", 2: "right_2_2", 3: "right_3_2"}
OPTION_COUNT = 3
SUPPORTED_PROMPT_TYPES = {"pure_audio", "no_audio_image", "audio_text", "audio_image", "no_audio_text"}
OPTION_NODE_NAMES = set(NODE_NAME_BY_OPTION.values())
PROMPT_TEXT_NODE_NAME = "题干文本"
PROMPT_TEXT_STYLE = {
    "x": 0,
    "y": 250,
    "w": 980,
    "h": 180,
    "fontSize": 72,
    "color": "#000000",
    "fontFamily": "FZCuYuan-M03S",
}


class ConfigError(RuntimeError):
    pass


def clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def load_questions(xlsx_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ConfigError(f"Sheet not found: {sheet_name}. Available sheets: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]

    questions: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = [clean_cell(v) for v in row]
        qno, audio, stem_img, stem_text, opt_no, opt_img, opt_text, is_correct, *_ = cells + [None] * 8
        if qno is None and opt_no is None:
            continue

        if qno is not None:
            current = {
                "question_no": int(qno),
                "audio_name": audio,
                "stem_img_name": stem_img,
                "stem_text": stem_text,
                "options": [],
                "source_row": row_index,
            }
            questions.append(current)

        if current is None:
            raise ConfigError(f"Row {row_index}: option row appears before a question number")
        if opt_no is None:
            raise ConfigError(f"Row {row_index}: missing option number")

        current["options"].append(
            {
                "option_no": int(opt_no),
                "option_img_name": opt_img,
                "option_text": opt_text,
                "is_correct": is_correct not in (None, "", 0, "0", "否", "false", "False"),
                "source_row": row_index,
            }
        )

    if not questions:
        raise ConfigError(f"No questions found in {xlsx_path}#{sheet_name}")
    return questions


def normalize_category(row: dict[str, Any]) -> str:
    category = (row.get("category") or row.get("type") or "").lower()
    url = row.get("url") or row.get("URL") or ""
    if not category:
        if "/audio/" in url or url.endswith((".mp3", ".wav", ".m4a")):
            category = "audio"
        elif "/image/" in url or url.endswith((".png", ".jpg", ".jpeg", ".webp")):
            category = "image"
    return category


def row_url(row: dict[str, Any]) -> str:
    return row.get("url") or row.get("URL") or ""


def load_resource_lookup(resource_path: Path, allow_duplicate_names: bool = False) -> dict[str, dict[str, Any]]:
    raw = json.loads(resource_path.read_text(encoding="utf-8"))
    rows = raw if isinstance(raw, list) else raw.get("rows", [])
    if not isinstance(rows, list):
        raise ConfigError(f"Resource file must be a list or contain rows[]: {resource_path}")

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        name = row.get("name") or row.get("名字")
        if name:
            grouped[str(name).strip()].append(row)

    duplicates = {name: values for name, values in grouped.items() if len(values) > 1}
    if duplicates and not allow_duplicate_names:
        names = ", ".join(sorted(duplicates)[:12])
        raise ConfigError(
            "Duplicate resource names found. Confirm reuse/rename before generation, "
            f"or pass --allow-duplicate-resource-names after confirmation. Names: {names}"
        )

    return {name: values[-1] for name, values in grouped.items()}


def require_resource(
    lookup: dict[str, dict[str, Any]],
    name: str | None,
    category: str,
    *,
    label: str,
) -> dict[str, Any] | None:
    if not name:
        return None
    row = lookup.get(name)
    if not row:
        raise ConfigError(f"Missing {category} resource for {label}: {name}")
    actual_category = normalize_category(row)
    if actual_category and actual_category != category:
        raise ConfigError(f"Resource category mismatch for {label}: {name} is {actual_category}, expected {category}")
    url = row_url(row)
    if not url:
        raise ConfigError(f"Resource URL is empty for {label}: {name}")
    return row


def get_state(component: dict[str, Any], state_name: str) -> dict[str, Any] | None:
    for state in component.get("component_data", {}).get("states", []):
        if state.get("state") == state_name:
            return state
    return None


def ensure_state(component: dict[str, Any], state_name: str, fallback_state_name: str = "default") -> dict[str, Any]:
    state = get_state(component, state_name)
    if state:
        return state

    states = component["component_data"].setdefault("states", [])
    fallback = deepcopy(get_state(component, fallback_state_name) or states[0])
    fallback["state"] = state_name
    labels = {
        "default": "默认",
        "clickStart": "按下",
        "clickEnd": "抬起",
        "compLoadFinish": "组件加载完成",
        "level_correct": "全局正确",
        "level_wrong": "全局错误",
    }
    fallback["label"] = labels.get(state_name, state_name)
    states.append(fallback)
    return fallback


def ensure_source_entry(state: dict[str, Any], source_name: str, fallback: dict[str, Any] | None = None) -> dict[str, Any]:
    source = state.setdefault("source", {})
    if source_name not in source:
        source[source_name] = deepcopy(fallback or {})
    return source[source_name]


def set_active_hide(state: dict[str, Any]) -> None:
    state.setdefault("active", {})
    state["active"]["canEdit"] = True
    state["active"]["switch"] = True
    state["active"]["value"] = "hide"


def state_transform(component: dict[str, Any], state_name: str = "default") -> dict[str, Any]:
    state = get_state(component, state_name) or {}
    return state.get("transform", {})


def source_value(component: dict[str, Any], source_name: str, state_name: str = "default") -> str:
    state = get_state(component, state_name) or {}
    return state.get("source", {}).get(source_name, {}).get("value", "") or ""


def has_source(component: dict[str, Any], source_name: str) -> bool:
    for state in component.get("component_data", {}).get("states", []):
        if source_name in state.get("source", {}):
            return True
    return False


def set_node_content(
    component: dict[str, Any],
    *,
    sprite_url: str | None,
    text_value: str | None,
    preserve_existing_sprite: bool = False,
) -> None:
    for state_name in ["default", "compLoadFinish", "level_correct"]:
        state = ensure_state(component, state_name, "default") if state_name == "compLoadFinish" else get_state(component, state_name)
        if not state:
            continue
        source = state.setdefault("source", {})
        if "MSprite" in source:
            if sprite_url is not None:
                source["MSprite"]["value"] = sprite_url
            elif not preserve_existing_sprite:
                source["MSprite"]["value"] = ""
        if "MLabel" in source:
            source["MLabel"]["value"] = text_value or ""


def set_label_content(component: dict[str, Any], text_value: str) -> None:
    for state_name in ["default", "compLoadFinish", "level_correct"]:
        state = ensure_state(component, state_name, "default") if state_name == "compLoadFinish" else get_state(component, state_name)
        if state:
            ensure_source_entry(state, "MLabel")["value"] = text_value


def set_sprite_content(component: dict[str, Any], sprite_url: str) -> None:
    for state_name in ["default", "compLoadFinish", "level_correct"]:
        state = ensure_state(component, state_name, "default") if state_name == "compLoadFinish" else get_state(component, state_name)
        if state:
            ensure_source_entry(state, "MSprite")["value"] = sprite_url


def remove_level_correct_state(component: dict[str, Any]) -> None:
    states = component["component_data"].get("states", [])
    component["component_data"]["states"] = [s for s in states if s.get("state") != "level_correct"]


def infer_question_type(question: dict[str, Any]) -> str:
    has_audio = bool(question.get("audio_name"))
    has_stem_img = bool(question.get("stem_img_name"))
    has_stem_text = bool(question.get("stem_text"))
    if has_audio and not has_stem_img and not has_stem_text:
        return "pure_audio"
    if not has_audio and has_stem_img:
        return "no_audio_image"
    if has_audio and has_stem_text and not has_stem_img:
        return "audio_text"
    if has_audio and has_stem_img:
        return "audio_image"
    if has_stem_text:
        return "no_audio_text"
    return "unknown_prompt"


def infer_option_type(question: dict[str, Any]) -> str:
    has_images = [bool(opt.get("option_img_name")) for opt in question["options"]]
    has_texts = [bool(opt.get("option_text")) for opt in question["options"]]
    if all(has_images) and not any(has_texts):
        return "image"
    if all(has_texts) and not any(has_images):
        return "text"
    if all(img or text for img, text in zip(has_images, has_texts)):
        return "mixed_or_layered"
    return "invalid"


def validate_questions(questions: list[dict[str, Any]], resource_lookup: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    meta_levels: list[dict[str, Any]] = []
    qnos = [q["question_no"] for q in questions]
    dup_qnos = [qno for qno, count in Counter(qnos).items() if count > 1]
    if dup_qnos:
        raise ConfigError(f"Duplicate question numbers: {dup_qnos}")

    for question in questions:
        qno = question["question_no"]
        options = question["options"]
        if len(options) != OPTION_COUNT:
            raise ConfigError(f"Question {qno}: expected {OPTION_COUNT} options, got {len(options)}")

        option_numbers = sorted(opt["option_no"] for opt in options)
        if option_numbers != [1, 2, 3]:
            raise ConfigError(f"Question {qno}: option numbers must be [1, 2, 3], got {option_numbers}")

        correct = [opt for opt in options if opt["is_correct"]]
        if len(correct) != 1:
            raise ConfigError(f"Question {qno}: expected exactly one correct option, got {len(correct)}")

        question_type = infer_question_type(question)
        option_type = infer_option_type(question)
        if question_type not in SUPPORTED_PROMPT_TYPES:
            raise ConfigError(f"Question {qno}: unsupported prompt type '{question_type}'")
        if option_type == "invalid":
            raise ConfigError(f"Question {qno}: every option must provide image or text")

        if question.get("audio_name"):
            require_resource(resource_lookup, question["audio_name"], "audio", label=f"question {qno} audio")
        if question.get("stem_img_name"):
            require_resource(resource_lookup, question["stem_img_name"], "image", label=f"question {qno} stem image")
        for opt in options:
            if opt.get("option_img_name"):
                require_resource(resource_lookup, opt["option_img_name"], "image", label=f"question {qno} option {opt['option_no']}")

        meta_levels.append(
            {
                "question_no": qno,
                "question_type": question_type,
                "option_type": option_type,
                "correct_option": correct[0]["option_no"],
            }
        )
    return meta_levels


def is_system_or_choice_component(comp: dict[str, Any]) -> bool:
    cd_name = comp.get("component_data", {}).get("name", "")
    comp_name = comp.get("name")
    return (
        comp_name in {"TitleStem", "LevelNumber", "AloneClickChoice"}
        or cd_name in OPTION_NODE_NAMES
        or cd_name in {"背景", "节点_102", "节点_106", "节点_107"}
    )


def option_text_nodes(level: dict[str, Any]) -> dict[int, dict[str, Any]]:
    candidates = []
    for comp in level.get("components", []):
        cd = comp.get("component_data", {})
        if is_system_or_choice_component(comp) or cd.get("base") != "MLabel" or not has_source(comp, "MLabel"):
            continue
        tr = state_transform(comp)
        x = tr.get("x")
        y = tr.get("y", 0)
        if x is None:
            continue
        # Option text blocks sit in the answer row. Prompt text is high in the center.
        if y < 150:
            candidates.append((x, comp))
    if len(candidates) < 3:
        return {}
    candidates = sorted(candidates, key=lambda item: item[0])[:3]
    return {idx: comp for idx, (_, comp) in enumerate(candidates, start=1)}


def prompt_text_node(level: dict[str, Any]) -> dict[str, Any] | None:
    candidates = []
    for comp in level.get("components", []):
        cd = comp.get("component_data", {})
        if is_system_or_choice_component(comp) or cd.get("base") != "MLabel" or not has_source(comp, "MLabel"):
            continue
        tr = state_transform(comp)
        x = tr.get("x", 9999)
        y = tr.get("y", -9999)
        # Audio+text prompt samples put the central prompt near the top center.
        if y >= 150:
            candidates.append((abs(x), -y, comp))
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: (item[0], item[1]))[0][2]


def max_component_index(level: dict[str, Any]) -> int:
    indexes = [comp.get("index") for comp in level.get("components", []) if isinstance(comp.get("index"), int)]
    return max(indexes, default=-1)


def create_prompt_text_component(text_value: str, level: dict[str, Any]) -> dict[str, Any]:
    """Create the central prompt text node used by no-audio pure text stems.

    The monster reference template does not always carry this node. Historical
    CoursewareMaker BaseComponent MLabel structure is reused, while size/color
    are fixed to the monster prompt text style documented in the workflow.
    """
    state = {
        "groupKey": "",
        "state": "default",
        "label": "默认",
        "notDelete": False,
        "transform": {
            "x": PROMPT_TEXT_STYLE["x"],
            "y": PROMPT_TEXT_STYLE["y"],
            "w": PROMPT_TEXT_STYLE["w"],
            "h": PROMPT_TEXT_STYLE["h"],
            "scaleX": 1,
            "scaleY": 1,
            "rot": 0,
            "editRot": False,
            "anchorX": 0.5,
            "anchorY": 0.5,
        },
        "source": {
            "MLabel": {
                "value": text_value,
                "color": PROMPT_TEXT_STYLE["color"],
                "fontFamily": PROMPT_TEXT_STYLE["fontFamily"],
                "fontSize": PROMPT_TEXT_STYLE["fontSize"],
                "isBold": False,
                "isItalic": False,
                "isUnderline": False,
                "alignType": "center",
                "interval": [0, 0, 0, 0, 0],
                "closeable": True,
            }
        },
        "jump": {"canEdit": True, "opened": 0, "type": "", "to": "", "duration": 0},
        "active": {"canEdit": True, "switch": False, "value": "show"},
    }
    comp = {
        "component_data": {
            "id": f"gamenext_component_uuid_{uuid4()}",
            "edit_description": "",
            "component_id": "",
            "name": PROMPT_TEXT_NODE_NAME,
            "zIndex": 13,
            "base": "MLabel",
            "components": {
                "tools": {},
                "source": {"MLabel": 1},
                "lockState": {"state": False, "componentId": "", "componentName": "", "componentState": ""},
                "judgeRules": {"forbidIfCorrect": False, "inAnswerPool": False},
                "webEditorCustomInfo": {"chain": "", "isAnswerComponent": False, "isJudgeComponent": False},
            },
            "edit": {"lock": True, "curState": "default", "baseNotChange": False},
            "custom": [],
            "state_group": [],
            "states": [state, deepcopy({**state, "state": "compLoadFinish", "label": "组件加载完成"})],
            "event": {
                "eventMap": [{"label": "点击", "name": "click", "sys": False}],
                "dispatchFunList": [
                    {"allowEventName": ["all"], "label": "显示", "name": "onShow"},
                    {"allowEventName": ["all"], "label": "隐藏", "name": "onHide"},
                    {"allowEventName": ["all"], "label": "旋转", "name": "onRotated"},
                ],
                "value": [],
            },
        },
        "component_id": "BaseComponent",
        "component_name": "节点",
        "component_url": "",
        "index": max_component_index(level) + 1,
        "name": "BaseComponent",
        "version": "0.0.0",
    }
    level.setdefault("components", []).append(comp)
    return comp


def ensure_prompt_text_node(level: dict[str, Any], text_value: str) -> dict[str, Any]:
    node = prompt_text_node(level)
    if node:
        return node
    return create_prompt_text_component(text_value, level)


def prompt_image_node(level: dict[str, Any]) -> dict[str, Any] | None:
    candidates = []
    for comp in level.get("components", []):
        if is_system_or_choice_component(comp) or not has_source(comp, "MSprite"):
            continue
        tr = state_transform(comp)
        x = tr.get("x", 9999)
        y = tr.get("y", -9999)
        w = tr.get("w", 0) or 0
        h = tr.get("h", 0) or 0
        value = source_value(comp, "MSprite")
        # Center prompt images in historical refs are around x=0, y=350, w≈300, h≈220.
        if value and abs(x) < 180 and y > 150 and 120 <= w <= 520 and 100 <= h <= 420:
            candidates.append((abs(x), -y, comp))
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: (item[0], item[1]))[0][2]


def index_level_components(level: dict[str, Any], question_type: str) -> dict[str, Any]:
    components = level["components"]
    comp_by_node_name: dict[str, dict[str, Any]] = {}
    click_by_position: dict[int, dict[str, Any]] = {}
    title_stem = None
    effect_monster = None
    level_number = None

    for comp in components:
        comp_name = comp.get("name")
        cd = comp.get("component_data", {})
        cd_name = cd.get("name", "")

        if comp_name == "TitleStem":
            title_stem = comp
        elif comp_name == "LevelNumber":
            level_number = comp
        elif comp_name == "AloneClickChoice":
            states = cd.get("states", [])
            x = states[0].get("transform", {}).get("x", 0) if states else 0
            if x < -200:
                click_by_position[1] = comp
            elif x < 200:
                click_by_position[2] = comp
            else:
                click_by_position[3] = comp
        elif cd_name in set(NODE_NAME_BY_OPTION.values()):
            comp_by_node_name[cd_name] = comp
        elif cd_name == "节点_102":
            effect_monster = comp

    missing = []
    if question_type not in {"no_audio_image", "no_audio_text"} and title_stem is None:
        missing.append("TitleStem")
    if level_number is None:
        missing.append("LevelNumber")
    if effect_monster is None:
        missing.append("节点_102")
    for option_no, node_name in NODE_NAME_BY_OPTION.items():
        if node_name not in comp_by_node_name:
            missing.append(node_name)
        if option_no not in click_by_position:
            missing.append(f"AloneClickChoice position {option_no}")
    if missing:
        raise ConfigError(f"Template level is missing required monster components: {', '.join(missing)}")

    return {
        "level": level,
        "nodes": comp_by_node_name,
        "clicks": click_by_position,
        "title_stem": title_stem,
        "effect_monster": effect_monster,
        "level_number": level_number,
        "option_text_nodes": option_text_nodes(level),
        "prompt_text_node": prompt_text_node(level),
        "prompt_image_node": prompt_image_node(level),
    }


def update_title_stem(title_stem: dict[str, Any], audio_url: str | None) -> None:
    click_end = ensure_state(title_stem, "clickEnd")
    click_end.setdefault("source", {}).setdefault("MAudio", {})
    click_end["source"]["MAudio"]["value"] = audio_url or ""
    click_end["source"]["MAudio"]["audioType"] = "play_audio" if audio_url else "play_effect_1"

    comp_load_finish = ensure_state(title_stem, "compLoadFinish")
    comp_load_finish.setdefault("source", {}).setdefault("MAudio", {})
    comp_load_finish["source"]["MAudio"]["value"] = ""
    comp_load_finish["jump"] = {
        "canEdit": True,
        "opened": 1,
        "type": "countdown",
        "to": "clickEnd",
        "duration": 0.5,
    }

    title_correct = ensure_state(title_stem, "level_correct", "default")
    set_active_hide(title_correct)


def update_prompt(
    indexed: dict[str, Any],
    question: dict[str, Any],
    resource_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    qno = question["question_no"]
    question_type = infer_question_type(question)
    prompt_meta: dict[str, Any] = {"question_type": question_type}

    if question_type == "pure_audio":
        return prompt_meta

    if question_type in {"audio_text", "no_audio_text"}:
        node = indexed["prompt_text_node"] or ensure_prompt_text_node(indexed["level"], question.get("stem_text") or "")
        set_label_content(node, question.get("stem_text") or "")
        prompt_meta.update(
            {
                "prompt_mode": "text",
                "prompt_node_name": node.get("component_data", {}).get("name"),
                "prompt_text": question.get("stem_text") or "",
                "prompt_style": PROMPT_TEXT_STYLE,
            }
        )
        return prompt_meta

    if question_type in {"audio_image", "no_audio_image"}:
        node = indexed["prompt_image_node"]
        if not node:
            raise ConfigError(f"Question {qno}: template has no central prompt image node")
        img_res = require_resource(resource_lookup, question.get("stem_img_name"), "image", label=f"question {qno} stem image")
        image_url = row_url(img_res) if img_res else ""
        set_sprite_content(node, image_url)
        prompt_meta.update(
            {
                "prompt_mode": "image",
                "prompt_node_name": node.get("component_data", {}).get("name"),
                "prompt_image_name": question.get("stem_img_name"),
                "prompt_image_url": image_url,
            }
        )
        return prompt_meta

    raise ConfigError(f"Question {qno}: unsupported prompt type '{question_type}'")


def update_level(level: dict[str, Any], question: dict[str, Any], resource_lookup: dict[str, dict[str, Any]], total_levels: int) -> dict[str, Any]:
    question_type = infer_question_type(question)
    indexed = index_level_components(level, question_type)
    nodes = indexed["nodes"]
    clicks = indexed["clicks"]
    title_stem = indexed["title_stem"]
    effect_monster = indexed["effect_monster"]
    level_number = indexed["level_number"]
    text_nodes = indexed["option_text_nodes"]

    qno = question["question_no"]
    audio_res = require_resource(resource_lookup, question.get("audio_name"), "audio", label=f"question {qno} audio")
    audio_url = row_url(audio_res) if audio_res else None
    correct_option = next(opt["option_no"] for opt in question["options"] if opt["is_correct"])

    if title_stem:
        update_title_stem(title_stem, audio_url)

    default_state = get_state(level_number, "default")
    if default_state and "MRichText" in default_state.get("source", {}):
        default_state["source"]["MRichText"]["value"] = f"{qno}/{total_levels}"

    option_meta = []
    for opt in question["options"]:
        option_no = opt["option_no"]
        node_name = NODE_NAME_BY_OPTION[option_no]
        node_comp = nodes[node_name]
        click_comp = clicks[option_no]

        img_res = require_resource(
            resource_lookup,
            opt.get("option_img_name"),
            "image",
            label=f"question {qno} option {option_no}",
        )
        sprite_url = row_url(img_res) if img_res else None
        text_value = opt.get("option_text") or ""

        preserve_sprite = sprite_url is None and bool(text_value)
        set_node_content(node_comp, sprite_url=sprite_url, text_value=text_value, preserve_existing_sprite=preserve_sprite)
        text_node = text_nodes.get(option_no)
        if text_node and text_value:
            set_label_content(text_node, text_value)
        click_comp["component_data"]["components"]["tools"]["AloneClickChoice"]["anwserConfig"]["anwserRadio"] = (
            1 if opt["is_correct"] else 2
        )

        if opt["is_correct"]:
            correct_state = ensure_state(node_comp, "level_correct", "default")
            set_node_content(node_comp, sprite_url=sprite_url, text_value=text_value, preserve_existing_sprite=preserve_sprite)
            set_active_hide(correct_state)
            if text_node:
                text_correct = ensure_state(text_node, "level_correct", "default")
                if text_value:
                    set_label_content(text_node, text_value)
                set_active_hide(text_correct)
        else:
            remove_level_correct_state(node_comp)
            if text_node:
                remove_level_correct_state(text_node)

        option_meta.append(
            {
                "option_no": option_no,
                "node_name": node_name,
                "text_node_name": text_node.get("component_data", {}).get("name") if text_node else None,
                "click_component": click_comp.get("component_data", {}).get("name"),
                "image_name": opt.get("option_img_name"),
                "image_url": sprite_url,
                "text": text_value,
                "is_correct": bool(opt["is_correct"]),
            }
        )

    effect_correct = ensure_state(effect_monster, "level_correct", "default")
    effect_correct.setdefault("source", {}).setdefault("MSpine", {})
    effect_correct["source"]["MSpine"]["animation"] = RIGHT_ANIMATION_BY_OPTION[correct_option]

    prompt_meta = update_prompt(indexed, question, resource_lookup)

    return {
        "question_no": qno,
        "question_type": question_type,
        "option_type": infer_option_type(question),
        "audio_name": question.get("audio_name"),
        "audio_url": audio_url,
        "stem_img_name": question.get("stem_img_name"),
        "stem_text": question.get("stem_text"),
        "prompt": prompt_meta,
        "correct_option": correct_option,
        "feedback_animation": RIGHT_ANIMATION_BY_OPTION[correct_option],
        "options": option_meta,
    }


def load_template(path: Path) -> dict[str, Any]:
    raw = path.read_text(encoding="utf-8").strip()
    if raw.startswith("'") and raw.endswith("'"):
        raw = raw[1:-1]
    data = json.loads(raw)
    if "result" in data and "configuration" in data["result"]:
        cfg = data["result"]["configuration"]
        return json.loads(cfg) if isinstance(cfg, str) else cfg
    if "configuration" in data and isinstance(data["configuration"], str):
        return json.loads(data["configuration"])
    if "game" in data:
        return data
    raise ConfigError(f"Template is not a recognized game config: {path}")


def build_config(args: argparse.Namespace) -> dict[str, Any]:
    questions = load_questions(args.xlsx, args.sheet)
    resource_lookup = load_resource_lookup(args.resources, args.allow_duplicate_resource_names)
    preflight = validate_questions(questions, resource_lookup)
    config = load_template(args.template)

    total_levels = len(questions)
    if len(config.get("game", [])) < total_levels:
        raise ConfigError(f"Template has {len(config.get('game', []))} levels, but {total_levels} questions were provided")
    config["game"] = deepcopy(config["game"][:total_levels])

    build_meta = {
        "schema": "coursewaremaker.monster.build_meta.v1",
        "sheet_name": args.sheet,
        "question_count": total_levels,
        "template": str(args.template),
        "resources": str(args.resources),
        "preflight": preflight,
        "levels": [],
    }
    for level, question in zip(config["game"], questions):
        build_meta["levels"].append(update_level(level, question, resource_lookup, total_levels))

    return {"config": config, "meta": build_meta}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_QUESTION_XLSX, help="Question workbook path")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Question sheet name")
    parser.add_argument("--resources", type=Path, default=DEFAULT_RESOURCE_JSON, help="Resource JSON path")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE_JSON, help="Monster reference/template JSON path")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_JSON, help="Output config JSON path")
    parser.add_argument("--meta", type=Path, default=DEFAULT_OUTPUT_META, help="Output build meta JSON path")
    parser.add_argument("--allow-duplicate-resource-names", action="store_true", help="Allow duplicate resource names after explicit user confirmation")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        result = build_config(args)
    except ConfigError as exc:
        print(f"ERROR: {exc}")
        return 2

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.meta.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result["config"], ensure_ascii=False, indent=2), encoding="utf-8")
    args.meta.write_text(json.dumps(result["meta"], ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result["meta"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
