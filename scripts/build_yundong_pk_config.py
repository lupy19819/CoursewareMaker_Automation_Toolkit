import json
import re
import sys
from copy import deepcopy
from pathlib import Path

import openpyxl


ROOT = Path(r"D:\codexProject")
RESOURCE_JSON = ROOT / "latest_resources.json"
QUESTION_XLSX = ROOT / "zhiyinlou_race_test_latest.xlsx"
RUN_STABLE_DETAIL = ROOT / "8fbe9ab9-28fa-11f1-906b-da4a8224db76.detail.json"
SWIM_ENV_DETAIL = ROOT / "1034b0ba-e534-11f0-9165-0e324dbd00ee.detail.json"
RACECAR_ENV_DETAIL = Path(r"D:\迅雷下载\赛车玩法_测试配置_单行版.json")


def resolve_sheet_name(workbook: openpyxl.Workbook, sheet_name: str) -> str:
    if sheet_name in workbook.sheetnames:
        return sheet_name
    for name in workbook.sheetnames:
        if name.startswith(sheet_name):
            return name
    for name in workbook.sheetnames:
        if sheet_name in name:
            return name
    raise KeyError(f"Worksheet not found: {sheet_name}")


def load_questions(xlsx_path: Path, sheet_name: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = resolve_sheet_name(wb, sheet_name)
    ws = wb[sheet_name]

    questions = []
    current = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        qno, audio, stem_img, opt_no, opt_img, opt_text, is_correct, *_ = row
        if qno is None and opt_no is None:
            continue

        if qno is not None:
            current = {
                "question_no": int(qno),
                "audio_name": audio,
                "stem_img_name": stem_img,
                "options": [],
            }
            questions.append(current)

        current["options"].append(
            {
                "option_no": int(opt_no),
                "option_img_name": opt_img,
                "option_text": opt_text,
                "is_correct": is_correct is not None,
            }
        )

    return questions


def load_resource_rows() -> list[dict]:
    return json.loads(RESOURCE_JSON.read_text(encoding="utf-8"))["rows"]


def load_resource_lookup(rows: list[dict]) -> dict[str, dict]:
    return {row["name"]: row for row in rows if row.get("name")}


def resolve_resource(name: str, category: str, lookup: dict[str, dict], rows: list[dict]) -> dict:
    if name in lookup:
        return lookup[name]

    suffix_match = re.search(r"([A-Za-z]+)$", name or "")
    suffix = suffix_match.group(1) if suffix_match else name.rsplit("音频", 1)[-1]
    candidates = [row for row in rows if row.get("category") == category and (row.get("name") or "").endswith(suffix)]
    if category == "image":
        preferred = [row for row in candidates if any(keyword in (row.get("name") or "") for keyword in ("游泳", "赛车", "赛跑"))]
        if preferred:
            candidates = preferred
    if len(candidates) == 1:
        return candidates[0]
    if candidates:
        return candidates[0]
    raise KeyError(f"Missing {category} resource: {name}")


def load_template_from_game_detail(path: Path) -> tuple[dict, dict]:
    detail = json.loads(path.read_text(encoding="utf-8"))
    if "result" in detail and "configuration" in detail["result"]:
        result = detail["result"]
        config = json.loads(result["configuration"])
        return result, config
    if "custom_game" in detail and "common" in detail:
        return {"game_name": path.stem, "configuration": None}, detail
    raise KeyError(f"Unsupported template format: {path}")


def infer_subtype(sheet_name: str) -> str:
    if "游泳" in sheet_name:
        return "swim"
    if "赛车" in sheet_name:
        return "racecar"
    return "run"


def apply_environment_template(config: dict, env_config: dict) -> None:
    custom = config["common"]["custom"]
    env_custom = env_config["common"]["custom"]
    custom["additional_environment"] = deepcopy(env_custom["additional_environment"])
    if "environment" in custom and "environment" in env_custom:
        custom["environment"] = deepcopy(env_custom["environment"])


def apply_topic_style_template(config: dict, style_config: dict) -> None:
    template_topic = style_config["custom_game"][0]["topics"][0]
    template_title_res = template_topic["title_res"]
    template_options = template_title_res["options"]

    for level in config["custom_game"]:
        topic = level["topics"][0]
        title_res = topic["title_res"]
        title_res["titleBg"] = deepcopy(template_title_res.get("titleBg", ""))
        if "audioSpine" in template_title_res:
            title_res["audioSpine"] = deepcopy(template_title_res["audioSpine"])

        for index, option in enumerate(title_res.get("options", [])):
            template_item = template_options[min(index, len(template_options) - 1)]["item"]
            item = option["item"]
            item["bgOptionNormal"] = deepcopy(template_item.get("bgOptionNormal", ""))
            item["bgOptionCorrect"] = deepcopy(template_item.get("bgOptionCorrect", ""))
            item["bgOptionWrong"] = deepcopy(template_item.get("bgOptionWrong", ""))
            if "correctSpine" in template_item:
                item["correctSpine"] = deepcopy(template_item["correctSpine"])


def calc_font_size(text: str, default: int) -> int:
    text = (text or "").strip()
    if not text:
        return default
    if len(text) <= 6:
        return default
    if len(text) <= 10:
        return max(52, default - 8)
    if len(text) <= 16:
        return max(44, default - 14)
    return max(36, default - 20)


def update_topic(topic: dict, question: dict, resource_lookup: dict[str, dict], resource_rows: list[dict]) -> dict:
    title_res = topic["title_res"]
    audio_res = resolve_resource(question["audio_name"], "audio", resource_lookup, resource_rows)

    title_res["btnAudio"] = audio_res["url"]
    title_res["titleAuido"] = audio_res["url"]

    stem_img_name = question.get("stem_img_name")
    if stem_img_name:
        stem_img_res = resource_lookup.get(stem_img_name)
        if not stem_img_res:
            raise KeyError(f"Missing stem image resource: {stem_img_name}")
        title_res["icon"] = stem_img_res["url"]
    else:
        title_res["icon"] = ""

    if isinstance(title_res.get("titleText"), dict):
        title_res["titleText"]["MLabel"] = ""

    if len(title_res["options"]) < len(question["options"]):
        template_option = deepcopy(title_res["options"][-1])
        while len(title_res["options"]) < len(question["options"]):
            title_res["options"].append(deepcopy(template_option))

    correct_option = None
    for index, option in enumerate(question["options"]):
        item = title_res["options"][index]["item"]
        option_img_name = option.get("option_img_name")
        option_text = (option.get("option_text") or "").strip()
        if option_img_name:
            option_img_res = resolve_resource(option_img_name, "image", resource_lookup, resource_rows)
            item["icon"] = option_img_res["url"]
            if isinstance(item.get("opstionText"), dict):
                item["opstionText"]["MLabel"] = ""
        else:
            item["icon"] = ""
            if isinstance(item.get("opstionText"), dict):
                item["opstionText"]["MLabel"] = option_text
                default_size = item["opstionText"].get("fontSize") or 70
                item["opstionText"]["fontSize"] = calc_font_size(option_text, default_size)
        item["switch"] = bool(option["is_correct"])
        if option["is_correct"]:
            correct_option = option["option_no"]

    return {
        "question_no": question["question_no"],
        "audio_name": question["audio_name"],
        "audio_url": audio_res["url"],
        "correct_option": correct_option,
        "option_type": "image" if question["options"][0].get("option_img_name") else "text",
    }


def main() -> None:
    if len(sys.argv) < 3:
        raise SystemExit("Usage: python build_yundong_pk_config.py <sheet_name> <game_detail_json> [question_xlsx]")

    sheet_name = sys.argv[1]
    game_detail_json = Path(sys.argv[2])
    question_xlsx = Path(sys.argv[3]) if len(sys.argv) > 3 else QUESTION_XLSX

    output_json = ROOT / f"{sheet_name}.config.json"
    output_meta = ROOT / f"{sheet_name}.build-meta.json"

    questions = load_questions(question_xlsx, sheet_name)
    target_result, _target_config = load_template_from_game_detail(game_detail_json)
    subtype = infer_subtype(sheet_name)
    skeleton_result, config = load_template_from_game_detail(RUN_STABLE_DETAIL)
    if subtype == "swim":
        _env_result, env_config = load_template_from_game_detail(SWIM_ENV_DETAIL)
        apply_environment_template(config, env_config)
        apply_topic_style_template(config, env_config)
    elif subtype == "racecar":
        _env_result, env_config = load_template_from_game_detail(RACECAR_ENV_DETAIL)
        apply_environment_template(config, env_config)
        apply_topic_style_template(config, env_config)

    resource_rows = load_resource_rows()
    resource_lookup = load_resource_lookup(resource_rows)

    total_levels = len(questions)
    config["custom_game"] = deepcopy(config["custom_game"][:total_levels])

    build_meta = {
        "sheet_name": sheet_name,
        "game_name": target_result["game_name"],
        "subtype": subtype,
        "skeleton_source": skeleton_result["game_name"],
        "question_count": total_levels,
        "levels": [],
    }

    for level, question in zip(config["custom_game"], questions, strict=True):
        topic = level["topics"][0]
        build_meta["levels"].append(update_topic(topic, question, resource_lookup, resource_rows))

    output_json.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    output_meta.write_text(json.dumps(build_meta, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(build_meta, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
