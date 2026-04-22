#!/usr/bin/env python3
"""Fetch CoursewareMaker resources and merge them into JSON/XLSX exports."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import os
import subprocess
import sys
import urllib.error
import urllib.request
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


DEFAULT_JSON = Path("/Users/tal/Documents/New project/编辑器上传素材_resources导出.json")
DEFAULT_XLSX = Path("/Users/tal/Documents/New project/编辑器上传素材_resources导出.xlsx")
API_URL = "https://sszt-gateway.speiyou.com/beibo/game/config/resources"
COURSEWARE_HOST = "coursewaremaker.speiyou.com"


def parse_int_list(value: str | None) -> list[int]:
    if not value:
        return []
    result: list[int] = []
    for part in value.replace(";", ",").split(","):
      part = part.strip()
      if not part:
          continue
      result.append(int(part))
    return result


def parse_update_time(value: str) -> str:
    return value or ""


def run_osascript(script: str) -> str:
    completed = subprocess.run(
        ["osascript", "-e", script],
        check=True,
        capture_output=True,
        text=True,
    )
    return completed.stdout.strip()


def find_chrome_token() -> str:
    script = r'''
tell application "Google Chrome"
  repeat with w in windows
    repeat with t in tabs of w
      set u to URL of t
      if u contains "coursewaremaker.speiyou.com" then
        return execute t javascript "localStorage.getItem('GAMEMAKER_TOKEN') || ''"
      end if
    end repeat
  end repeat
end tell
return ""
'''
    try:
        return run_osascript(script).strip()
    except Exception:
        return ""


def api_request(token: str, payload: dict) -> dict:
    req = urllib.request.Request(
        API_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json;charset=UTF-8",
            "beibotoken": token,
        },
        method="POST",
    )
    try:
      with urllib.request.urlopen(req, timeout=60) as resp:
            return {
                "status": resp.status,
                "body": resp.read().decode("utf-8", errors="replace"),
            }
    except urllib.error.HTTPError as exc:
        return {
            "status": exc.code,
            "body": exc.read().decode("utf-8", errors="replace"),
        }


def normalize_row(row: dict) -> dict:
    return {
        "id": row.get("id"),
        "name": row.get("name", "") or "",
        "category": row.get("category", "") or "",
        "url": row.get("url", "") or "",
        "desc": row.get("desc", "") or "",
        "topic": row.get("topic", "") or "",
        "tag": row.get("tag", "") or "",
        "type": row.get("type"),
        "creator_name": row.get("creator_name", "") or "",
        "update_time": row.get("update_time", "") or "",
    }


def merge_key(row: dict) -> str:
    if row.get("id") is not None:
        return str(row["id"])
    return f'{row.get("category", "")}||{row.get("name", "")}||{row.get("url", "")}'


def merge_rows(existing: list[dict], fresh: list[dict]) -> list[dict]:
    merged: "OrderedDict[str, dict]" = OrderedDict()
    for row in existing:
        merged[merge_key(row)] = row
    for row in fresh:
        key = merge_key(row)
        if key not in merged:
            merged[key] = row
            continue
        current = merged[key]
        merged[key] = row if parse_update_time(row.get("update_time", "")) >= parse_update_time(current.get("update_time", "")) else current

    rows = list(merged.values())
    rows.sort(
        key=lambda r: (
            r.get("update_time", ""),
            r.get("id") if r.get("id") is not None else -1,
        ),
        reverse=True,
    )
    return rows


def load_existing_rows(json_path: Path, xlsx_path: Path) -> list[dict]:
    if json_path.exists():
        raw = json.loads(json_path.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            return [normalize_row(item) for item in raw]
        if isinstance(raw, dict):
            rows = raw.get("rows", [])
            if isinstance(rows, list):
                return [normalize_row(item) for item in rows]

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        rows: list[dict] = []
        for ws in wb.worksheets:
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for values in ws.iter_rows(min_row=2, values_only=True):
                row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
                rows.append(normalize_row(row))
        return rows

    return []


def fetch_latest_rows(token: str, payload_base: dict, page_size: int) -> tuple[list[dict], int]:
    first = api_request(token, payload_base)
    if first["status"] != 200:
        raise RuntimeError(f"resources request failed: HTTP {first['status']}")

    data = json.loads(first["body"] or "{}")
    result = data.get("result") or {}
    total = int(result.get("total") or len(result.get("list") or []))
    rows = list(result.get("list") or [])
    total_pages = max(1, math.ceil(total / page_size))

    for page in range(2, total_pages + 1):
        payload = dict(payload_base)
        payload["page"] = page
        response = api_request(token, payload)
        if response["status"] != 200:
            continue
        page_data = json.loads(response["body"] or "{}")
        page_rows = ((page_data.get("result") or {}).get("list") or [])
        rows.extend(page_rows)

    return [normalize_row(row) for row in rows], total


def write_json(path: Path, payload: dict, rows: list[dict]) -> None:
    path.write_text(
        json.dumps(
            {
                "fetched_at": dt.datetime.now().astimezone().isoformat(),
                "payload": payload,
                "total": len(rows),
                "rows": rows,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def sheet_name_for(category: str) -> str:
    name = category or "other"
    return name[:31]


def write_xlsx(path: Path, rows: list[dict], split_by_category: bool) -> None:
    columns = [
        ("id", "id", 10),
        ("名字", "name", 26),
        ("category", "category", 14),
        ("URL", "url", 90),
        ("描述", "desc", 18),
        ("topic", "topic", 12),
        ("tag", "tag", 12),
        ("type", "type", 10),
        ("创建人", "creator_name", 12),
        ("更新时间", "update_time", 22),
    ]

    wb = Workbook()
    wb.remove(wb.active)

    groups: dict[str, list[dict]]
    if split_by_category:
        groups = {"image": [], "audio": [], "spine": [], "video": [], "other": []}
        for row in rows:
            groups.setdefault(row.get("category") or "other", []).append(row)
    else:
        groups = {"resources": rows}

    for sheet_name, sheet_rows in groups.items():
        ws = wb.create_sheet(title=sheet_name_for(sheet_name))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{chr(64 + len(columns))}1"
        for col_idx, (header, key, width) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = width
        for row_idx, row in enumerate(sheet_rows, start=2):
            for col_idx, (_, key, _) in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Fetch and merge CoursewareMaker resources.")
    parser.add_argument("--token", help="beibotoken value; defaults to Chrome localStorage")
    parser.add_argument("--input-json", default=str(DEFAULT_JSON))
    parser.add_argument("--input-xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--output-json", default=str(DEFAULT_JSON))
    parser.add_argument("--output-xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--page-size", type=int, default=500)
    parser.add_argument("--name", default="")
    parser.add_argument("--type", type=int, default=1)
    parser.add_argument("--platform", type=int, default=1)
    parser.add_argument("--topic", default="")
    parser.add_argument("--tag", default="")
    parser.add_argument("--subject-ids", default="")
    parser.add_argument("--split-by-category", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    token = args.token or os.environ.get("GAMEMAKER_TOKEN") or find_chrome_token()
    if not token:
        print("No beibotoken found. Open a CoursewareMaker tab or pass --token.", file=sys.stderr)
        return 1

    payload = {
        "page": 1,
        "page_size": args.page_size,
        "name": args.name,
        "type": args.type,
        "platform": args.platform,
        "topic": parse_int_list(args.topic),
        "tag": int(args.tag) if str(args.tag).strip() else None,
        "subject_ids": parse_int_list(args.subject_ids),
    }

    existing = load_existing_rows(Path(args.input_json), Path(args.input_xlsx))
    latest, fetched_total = fetch_latest_rows(token, payload, args.page_size)
    merged = merge_rows(existing, latest)

    summary = {
        "fetched_total": fetched_total,
        "fetched_rows": len(latest),
        "existing_rows": len(existing),
        "merged_rows": len(merged),
        "added_rows": len(merged) - len(existing),
        "counts": {},
    }
    for row in merged:
        category = row.get("category") or "other"
        summary["counts"][category] = summary["counts"].get(category, 0) + 1

    if not args.dry_run:
        write_json(Path(args.output_json), payload, merged)
        write_xlsx(Path(args.output_xlsx), merged, args.split_by_category)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
