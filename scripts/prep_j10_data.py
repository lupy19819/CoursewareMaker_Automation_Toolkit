"""
准备 J10 数据：
1. 更新 resources/latest_resources.json，加入J10资源
2. 创建 zhiyinlou_monster_test_j10.xlsx（贪吃J10）
3. 在 zhiyinlou_race_test_latest.xlsx 中追加 J10赛跑 sheet
"""
import json
from pathlib import Path
import openpyxl
from openpyxl import Workbook
import datetime

BASE = Path(__file__).resolve().parents[1]

# ── 1. 资源ID映射 ──────────────────────────────────────────────
J10_RESOURCES = [
    # 贪吃图片
    {"id": 75820, "name": "26暑国际小班10贪吃e",   "category": "image", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/image/345733/2026-05-07/6f6bd48f255c90c849b56054cfe251cb.png"},
    {"id": 75821, "name": "26暑国际小班10贪吃f",   "category": "image", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/image/345733/2026-05-07/a6ab25cac1b833e514bd229507fdff06.png"},
    {"id": 75822, "name": "26暑国际小班10贪吃g",   "category": "image", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/image/345733/2026-05-07/ea7dc037963c2528cd772cf4cee54322.png"},
    {"id": 75823, "name": "26暑国际小班10贪吃h",   "category": "image", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/image/345733/2026-05-07/616038a477e36a03786e13dc5c2e2aec.png"},
    # 贪吃音频
    {"id": 75824, "name": "26暑国际小班10贪吃音频e", "category": "audio", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/audio/345733/2026-05-07/c5bfa6c41d43dd3340efcf7ca9586673.mp3"},
    {"id": 75825, "name": "26暑国际小班10贪吃音频f", "category": "audio", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/audio/345733/2026-05-07/2f90546e50cb4f612c9922a2b59ec914.mp3"},
    {"id": 75826, "name": "26暑国际小班10贪吃音频g", "category": "audio", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/audio/345733/2026-05-07/c4cd41355554d003b4146867a36e0aca.mp3"},
    {"id": 75827, "name": "26暑国际小班10贪吃音频h", "category": "audio", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/audio/345733/2026-05-07/095421ce24170a713dc4cba729f1fef1.mp3"},
    # 赛跑图片
    {"id": 75828, "name": "26暑国际小班10赛跑bag",  "category": "image", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/image/345733/2026-05-07/8c127bf68c564b31a843d3582b447efa.png"},
    {"id": 75829, "name": "26暑国际小班10赛跑book", "category": "image", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/image/345733/2026-05-07/1452c4cead4b3eba85ea424f3479b4c6.png"},
    {"id": 75830, "name": "26暑国际小班10赛跑toy",  "category": "image", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/image/345733/2026-05-07/75551656ec57e4dc0c42c2e51ecb3fc7.png"},
    # 赛跑音频
    {"id": 75831, "name": "26暑国际小班10赛跑音频单词bag",  "category": "audio", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/audio/345733/2026-05-07/1c3317d9966a36a6462e4464d8494cb2.mp3"},
    {"id": 75832, "name": "26暑国际小班10赛跑音频单词book", "category": "audio", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/audio/345733/2026-05-07/c542ec22c05bf2f2baabc8fa3b7be0b8.mp3"},
    {"id": 75833, "name": "26暑国际小班10赛跑音频单词toy",  "category": "audio", "url": "https://courseware-maker-1252161091.cos.ap-beijing.myqcloud.com/assets/audio/345733/2026-05-07/f89b9ce4e24200ff40946a4d0929dd2a.mp3"},
]

# ── 2. 更新 resources/latest_resources.json ────────────────────
res_path = BASE / "resources" / "latest_resources.json"
res_path.parent.mkdir(parents=True, exist_ok=True)
existing = json.loads(res_path.read_text(encoding="utf-8")) if res_path.exists() else {"rows": []}
existing_ids = {r["id"] for r in existing["rows"] if r.get("id")}
ts = datetime.datetime.now().isoformat()
for r in J10_RESOURCES:
    if r["id"] not in existing_ids:
        existing["rows"].append({
            "id": r["id"], "name": r["name"], "category": r["category"],
            "url": r["url"], "desc": "", "topic": [1], "tag": [7],
            "type": 2 if r["category"] == "audio" else 1,
            "creator_name": "江昊", "update_time": ts
        })
res_path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"✓ resources/latest_resources.json 已更新，共{len(existing['rows'])}条")

# ── 3. 创建贪吃J10 XLSX ──────────────────────────────────────────
MONSTER_SHEET = "国际新小班暑J10贪吃小怪兽"
monster_questions = [
    # qno, audio, stem_img, stem_text, opt_no, opt_img, opt_text, is_correct
    (1, "26暑国际小班10贪吃音频e", None, None, 1, "26暑国际小班10贪吃e", None, "是"),
    (None, None, None, None, 2, "26暑国际小班10贪吃g", None, None),
    (None, None, None, None, 3, "26暑国际小班10贪吃f", None, None),
    (2, "26暑国际小班10贪吃音频f", None, None, 1, "26暑国际小班10贪吃g", None, None),
    (None, None, None, None, 2, "26暑国际小班10贪吃e", None, None),
    (None, None, None, None, 3, "26暑国际小班10贪吃f", None, "是"),
    (3, "26暑国际小班10贪吃音频g", None, None, 1, "26暑国际小班10贪吃f", None, None),
    (None, None, None, None, 2, "26暑国际小班10贪吃g", None, "是"),
    (None, None, None, None, 3, "26暑国际小班10贪吃h", None, None),
    (4, "26暑国际小班10贪吃音频h", None, None, 1, "26暑国际小班10贪吃g", None, None),
    (None, None, None, None, 2, "26暑国际小班10贪吃f", None, None),
    (None, None, None, None, 3, "26暑国际小班10贪吃h", None, "是"),
]
monster_xlsx = BASE / "zhiyinlou_monster_test_j10.xlsx"
wb_m = Workbook()
ws_m = wb_m.active
ws_m.title = MONSTER_SHEET
ws_m.append(["题目序号", "音频命名", "题干图片", "题干文本", "选项序号", "选项图片名称", "选项文本内容", "是否正确"])
for row in monster_questions:
    ws_m.append(row)
wb_m.save(monster_xlsx)
print(f"✓ {monster_xlsx} 已创建 ({len(monster_questions)}行)")

# ── 4. 在赛跑XLSX中添加J10赛跑sheet ─────────────────────────────
RACE_SHEET = "国际新小班暑J10跑酷赛跑"
race_questions = [
    # qno, audio, stem_img, opt_no, opt_img, opt_text, is_correct
    (1, "26暑国际小班10赛跑音频单词book", None, 1, "26暑国际小班10赛跑book", None, "是"),
    (None, None, None, 2, "26暑国际小班10赛跑bag", None, None),
    (None, None, None, 3, "26暑国际小班10赛跑toy", None, None),
    (2, "26暑国际小班10赛跑音频单词bag", None, 1, "26暑国际小班10赛跑toy", None, None),
    (None, None, None, 2, "26暑国际小班10赛跑bag", None, "是"),
    (None, None, None, 3, "26暑国际小班10赛跑book", None, None),
    (3, "26暑国际小班10赛跑音频单词toy", None, 1, "26暑国际小班10赛跑bag", None, None),
    (None, None, None, 2, "26暑国际小班10赛跑book", None, None),
    (None, None, None, 3, "26暑国际小班10赛跑toy", None, "是"),
    (4, "26暑国际小班10赛跑音频单词toy", None, 1, "26暑国际小班10赛跑toy", None, "是"),
    (None, None, None, 2, "26暑国际小班10赛跑book", None, None),
    (None, None, None, 3, "26暑国际小班10赛跑bag", None, None),
    (5, "26暑国际小班10赛跑音频单词book", None, 1, "26暑国际小班10赛跑book", None, "是"),
    (None, None, None, 2, "26暑国际小班10赛跑bag", None, None),
    (None, None, None, 3, "26暑国际小班10赛跑toy", None, None),
    (6, "26暑国际小班10赛跑音频单词bag", None, 1, "26暑国际小班10赛跑bag", None, "是"),
    (None, None, None, 2, "26暑国际小班10赛跑book", None, None),
    (None, None, None, 3, "26暑国际小班10赛跑toy", None, None),
    (7, "26暑国际小班10赛跑音频单词book", None, 1, "26暑国际小班10赛跑toy", None, None),
    (None, None, None, 2, "26暑国际小班10赛跑bag", None, None),
    (None, None, None, 3, "26暑国际小班10赛跑book", None, "是"),
    (8, "26暑国际小班10赛跑音频单词toy", None, 1, "26暑国际小班10赛跑bag", None, None),
    (None, None, None, 2, "26暑国际小班10赛跑toy", None, "是"),
    (None, None, None, 3, "26暑国际小班10赛跑book", None, None),
    (9, "26暑国际小班10赛跑音频单词bag", None, 1, "26暑国际小班10赛跑bag", None, "是"),
    (None, None, None, 2, "26暑国际小班10赛跑book", None, None),
    (None, None, None, 3, "26暑国际小班10赛跑toy", None, None),
    (10, "26暑国际小班10赛跑音频单词book", None, 1, "26暑国际小班10赛跑toy", None, None),
    (None, None, None, 2, "26暑国际小班10赛跑book", None, "是"),
    (None, None, None, 3, "26暑国际小班10赛跑bag", None, None),
]
race_xlsx = BASE / "zhiyinlou_race_test_latest.xlsx"
wb_r = openpyxl.load_workbook(race_xlsx)
if RACE_SHEET in wb_r.sheetnames:
    del wb_r[RACE_SHEET]
ws_r = wb_r.create_sheet(RACE_SHEET)
ws_r.append(["题目序号", "音频命名", "题干图片", "选项序号", "选项图片名称", "选项文本内容", "是否正确"])
for row in race_questions:
    ws_r.append(row)
wb_r.save(race_xlsx)
print(f"✓ {race_xlsx} 已追加sheet '{RACE_SHEET}' ({len(race_questions)}行)")
print("\n准备完成！可以运行配置生成脚本。")
