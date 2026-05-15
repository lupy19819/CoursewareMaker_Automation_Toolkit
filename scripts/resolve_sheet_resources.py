#!/usr/bin/env python3
"""Resolve resource names used by a question sheet and export a filtered list.

This intentionally checks only resources referenced by the current task sheet.
Historical duplicate names elsewhere in resources/latest_resources.json should
not block an unrelated generation run.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RESOURCES = ROOT / "resources" / "latest_resources.json"

HEADER_ALIASES = {
    "audio": {"音频命名", "题目音频", "题干音频", "audio", "audio_name"},
    "image": {"题干图片", "选项图片名称", "选项图片", "图片名称", "image", "image_name", "option_image", "stem_image"},
}


class ResourceResolveError(RuntimeError):
    pass


def clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return str(value).strip() or None


def normalize_category(row: dict[str, Any]) -> str:
    category = str(row.get("category") or row.get("type") or "").lower()
    url = str(row.get("url") or row.get("URL") or "")
    if category:
        return category
    if "/audio/" in url or url.lower().endswith((".mp3", ".wav", ".m4a")):
        return "audio"
    if "/image/" in url or url.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
        return "image"
    return ""


def resource_name(row: dict[str, Any]) -> str | None:
    return clean(row.get("name") or row.get("名字"))


def load_resource_rows(path: Path) -> list[dict[str, Any]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    rows = raw if isinstance(raw, list) else raw.get("rows", [])
    if not isinstance(rows, list):
        raise ResourceResolveError(f"Resource file must be a list or contain rows[]: {path}")
    return rows


def header_indexes(ws: Any) -> dict[int, str]:
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first:
        raise ResourceResolveError("Sheet has no header row")
    indexes: dict[int, str] = {}
    for idx, value in enumerate(first):
        name = clean(value)
        if not name:
            continue
        for category, aliases in HEADER_ALIASES.items():
            if name in aliases:
                indexes[idx] = category
    if not indexes:
        raise ResourceResolveError("No resource columns detected in sheet headers")
    return indexes


def collect_needed_resources(xlsx: Path, sheet: str) -> dict[str, str]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ResourceResolveError(f"Sheet not found: {sheet}. Available sheets: {', '.join(wb.sheetnames)}")
    ws = wb[sheet]
    indexes = header_indexes(ws)
    needed: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        for idx, category in indexes.items():
            value = clean(row[idx] if idx < len(row) else None)
            if value:
                existing = needed.get(value)
                if existing and existing != category:
                    raise ResourceResolveError(f"Resource used as both {existing} and {category}: {value}")
                needed[value] = category
    if not needed:
        raise ResourceResolveError(f"No resources referenced by {xlsx}#{sheet}")
    return needed


def resolve_resources(needed: dict[str, str], rows: list[dict[str, Any]], allow_duplicates: bool) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        name = resource_name(row)
        if name in needed:
            grouped[name].append(row)

    missing = sorted(set(needed) - set(grouped))
    duplicates = {name: values for name, values in grouped.items() if len(values) > 1}
    mismatches = []
    selected = []
    for name in sorted(grouped):
        expected = needed[name]
        values = grouped[name]
        for row in values:
            actual = normalize_category(row)
            if actual and actual != expected:
                mismatches.append({"name": name, "expected": expected, "actual": actual})
        selected.append(values[-1])

    if missing:
        raise ResourceResolveError("Missing resources: " + ", ".join(missing))
    if mismatches:
        raise ResourceResolveError("Resource category mismatch: " + json.dumps(mismatches, ensure_ascii=False))
    if duplicates and not allow_duplicates:
        names = ", ".join(sorted(duplicates))
        raise ResourceResolveError(
            "Duplicate names among resources used by this sheet. Confirm reuse/rename or pass --allow-duplicates. Names: "
            + names
        )

    manifest = {
        "needed_count": len(needed),
        "selected_count": len(selected),
        "duplicates": {name: len(values) for name, values in duplicates.items()},
        "resources": [{"name": name, "category": needed[name]} for name in sorted(needed)],
    }
    return selected, manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--resources", type=Path, default=DEFAULT_RESOURCES)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--allow-duplicates", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    needed = collect_needed_resources(args.xlsx, args.sheet)
    selected, manifest = resolve_resources(needed, load_resource_rows(args.resources), args.allow_duplicates)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(selected, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.manifest:
        args.manifest.parent.mkdir(parents=True, exist_ok=True)
        args.manifest.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({**manifest, "output": str(args.output), "manifest": str(args.manifest) if args.manifest else None}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
